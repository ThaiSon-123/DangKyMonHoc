"""Import 1 file xlsx chứa Chương trình đào tạo (CTĐT) vào DB.

Định dạng xlsx mong đợi (theo mẫu user):
  Header row: STT | Mã MH | Tên môn học | Số tín chỉ | Môn bắt buộc | Đã học |
              Môn học đã học và đạt | lý thuyết | thực hành
  Section row: "Học kỳ X - Năm học YYYY - YYYY" trong cột Tên môn học, các cột khác trống
  Data row:    1 | LING022 | Cơ sở lập trình (3+0) | 3 | x | x | x | 45 | 0

Usage:
  docker compose exec backend python manage.py import_curriculum_xlsx \\
      data/curriculums/CNTT-2023.xlsx \\
      --major CNTT --cohort-year 2023 \\
      --curriculum-code CNTT-2023 \\
      --curriculum-name "CTĐT Công nghệ thông tin khóa 2023"
"""
from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.courses.models import Course
from apps.curriculums.models import Curriculum, CurriculumCourse
from apps.majors.models import Major

SECTION_RE = re.compile(
    r"H[ọo]c k[ỳy]\s*(\d+)\s*-\s*N[ăa]m\s*h[ọo]c\s*(\d{4})\s*-\s*(\d{4})",
    re.IGNORECASE,
)


class Command(BaseCommand):
    help = "Import 1 file xlsx CTĐT vào DB."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Đường dẫn file xlsx (tương đối /app).")
        parser.add_argument("--major", required=True, help="Mã ngành (vd: CNTT).")
        parser.add_argument(
            "--cohort-year", type=int, required=True,
            help="Năm tuyển sinh của khóa (vd: 2023).",
        )
        parser.add_argument(
            "--curriculum-code", required=True,
            help="Mã CTĐT (vd: CNTT-2023).",
        )
        parser.add_argument("--curriculum-name", default="", help="Tên CTĐT.")
        parser.add_argument(
            "--total-credits", type=int, default=145,
            help="Tổng số tín chỉ yêu cầu của CTĐT (default 145).",
        )
        parser.add_argument(
            "--knowledge-block", default="MAJOR",
            choices=["GENERAL", "BASIC", "MAJOR", "ELECTIVE", "THESIS"],
            help="Default knowledge_block cho các môn (default MAJOR).",
        )
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("Cài openpyxl: pip install openpyxl")

        path = Path(opts["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Không tìm thấy file: {path}")

        major = Major.objects.filter(code=opts["major"]).first()
        if not major:
            raise CommandError(
                f"Ngành '{opts['major']}' chưa có. Tạo trước qua /admin/majors hoặc Django admin."
            )

        wb = load_workbook(filename=str(path), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("File rỗng.")

        # Parse rows
        current_semester_idx = 0  # 1, 2, 3, ...
        academic_year_seen: list[tuple[int, int]] = []  # tracking để tính HK liên tiếp
        parsed: list[dict] = []

        for row in rows[1:]:  # skip header
            if not row or all(v is None or v == "" for v in row):
                continue
            # Section header: "Học kỳ X - Năm học YYYY - YYYY"
            # Có thể nằm ở cột bất kỳ vì layout xlsx khác nhau
            row_text = " ".join(str(c) for c in row if c is not None)
            m = SECTION_RE.search(row_text)
            if m:
                term_no = int(m.group(1))
                year_start = int(m.group(2))
                key = (year_start, term_no)
                if key not in academic_year_seen:
                    academic_year_seen.append(key)
                # Map (year_start, term) → semester_idx liên tục
                current_semester_idx = academic_year_seen.index(key) + 1
                continue

            # Data row — yêu cầu có code (cột 2) và name (cột 3)
            try:
                stt, code, name, credits, required, _learned, _passed, theory, practice = row[:9]
            except ValueError:
                continue
            if not code or not isinstance(code, str):
                continue
            code = code.strip()
            if not code or len(code) > 16:
                continue

            # Strip "(LT+TH)" suffix khỏi tên
            name_clean = re.sub(r"\s*\(\d+\+\d+\)\s*$", "", str(name or "")).strip()
            try:
                credits = int(credits or 0)
            except (TypeError, ValueError):
                credits = 0
            try:
                theory = int(theory or 0)
            except (TypeError, ValueError):
                theory = 0
            try:
                practice = int(practice or 0)
            except (TypeError, ValueError):
                practice = 0
            is_required = bool(required) and str(required).strip().lower() in ("x", "true", "1", "có")

            parsed.append({
                "code": code,
                "name": name_clean,
                "credits": max(1, credits),
                "theory_hours": theory,
                "practice_hours": practice,
                "is_required": is_required,
                "suggested_semester": current_semester_idx,
            })

        if not parsed:
            raise CommandError("Không parse được môn nào trong file.")

        self.stdout.write(self.style.NOTICE(
            f"Parsed {len(parsed)} môn từ {path.name}, "
            f"trải {max(p['suggested_semester'] for p in parsed)} học kỳ."
        ))

        if opts["dry_run"]:
            for p in parsed[:5]:
                self.stdout.write(f"  HK{p['suggested_semester']:>2} {p['code']:8} {p['name']}")
            self.stdout.write(self.style.WARNING("[DRY-RUN] Không ghi DB."))
            return

        # Apply
        with transaction.atomic():
            curriculum, c_created = Curriculum.objects.update_or_create(
                code=opts["curriculum_code"],
                defaults={
                    "major": major,
                    "name": opts["curriculum_name"] or f"CTĐT {major.code} K{opts['cohort_year']}",
                    "cohort_year": opts["cohort_year"],
                    "total_credits_required": opts["total_credits"],
                    "is_active": True,
                },
            )
            self.stdout.write(self.style.SUCCESS(
                f"{'+ Tạo' if c_created else '~ Cập nhật'} Curriculum: {curriculum.code}"
            ))

            new_courses = 0
            updated_courses = 0
            new_links = 0
            updated_links = 0

            for p in parsed:
                course, created = Course.objects.update_or_create(
                    code=p["code"],
                    defaults={
                        "name": p["name"],
                        "credits": p["credits"],
                        "theory_hours": p["theory_hours"],
                        "practice_hours": p["practice_hours"],
                        "is_active": True,
                    },
                )
                if created:
                    new_courses += 1
                else:
                    updated_courses += 1

                _, link_created = CurriculumCourse.objects.update_or_create(
                    curriculum=curriculum, course=course,
                    defaults={
                        "is_required": p["is_required"],
                        "suggested_semester": p["suggested_semester"],
                        "knowledge_block": (
                            CurriculumCourse.Knowledge.GENERAL
                            if course.code.upper().startswith("KTCH")
                            else opts["knowledge_block"]
                        ),
                    },
                )
                if link_created:
                    new_links += 1
                else:
                    updated_links += 1

        self.stdout.write(self.style.SUCCESS(
            f"✓ Courses: +{new_courses} mới, ~{updated_courses} cập nhật"
        ))
        self.stdout.write(self.style.SUCCESS(
            f"✓ CurriculumCourse: +{new_links} mới, ~{updated_links} cập nhật"
        ))
